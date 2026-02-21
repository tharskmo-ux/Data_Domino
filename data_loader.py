import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataLoader:
    def __init__(self, config=None):
        """Initialize loader"""
        self.config = config or {}
        
    def load(self, df: pd.DataFrame, mapping_info: dict = None, 
             output_path: str = None) -> str:
        """
        Main method: Generate Excel file
        
        Args:
            df: Procurement data DataFrame
            mapping_info: Optional dict of field mappings
            output_path: Optional output path (auto-generated if None)
            
        Returns:
            Path to generated Excel file
        """
        try:
            # 1. Column Mapping
            df = df.copy() # Avoid modifying original DF
            if mapping_info:
                # Map source columns to internal names
                # Reverse mapping: source -> internal
                inv_map = {v: k for k, v in mapping_info.items() if v in df.columns}
                df = df.rename(columns=inv_map)

            # 2. Validation
            if df.empty:
                raise ValueError("Cannot generate report from empty DataFrame")
            
            # Ensure required internal fields exist
            required_fields = ['vendor', 'amount', 'date']
            missing = [f for f in required_fields if f not in df.columns]
            if missing:
                raise ValueError(f"Missing required fields: {missing}")

            # 3. Data Type Normalization
            # Amount should be numeric
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce')

            # Date Parsing Enhancement
            df['date'] = df['date'].apply(self._parse_date)

            # 4. Setup Output Path
            if output_path is None:
                timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
                output_path = f"procurement_analysis_{timestamp}.xlsx"
            
            # Ensure output directory exists if path contains directory
            output_dir = Path(output_path).parent
            if output_dir.name and not output_dir.exists():
                output_dir.mkdir(parents=True, exist_ok=True)

            # 5. Create Workbook
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 6. Generate Sheets
            logger.info("Generating 'Executive Summary'...")
            self._create_executive_summary(wb, df)
            
            logger.info("Generating 'Spend by Vendor'...")
            self._create_vendor_analysis(wb, df)
            
            logger.info("Generating 'Spend by Category'...")
            self._create_category_analysis(wb, df)
            
            logger.info("Generating 'Monthly Trends'...")
            self._create_monthly_trends(wb, df)
            
            logger.info("Generating 'Top Insights'...")
            self._create_insights(wb, df)
            
            logger.info("Generating 'Detailed Data'...")
            self._create_detailed_data(wb, df)
            
            logger.info("Generating 'Data Quality Report'...")
            self._create_data_quality_report(wb, df)
            
            # 7. Save
            wb.save(output_path)
            logger.info(f"Report saved successfully to: {output_path}")
            
            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}")
            raise

    def _parse_date(self, val):
        """Helper to parse dates including Excel serial numbers"""
        if pd.isna(val):
            return val
        if isinstance(val, (int, float, np.integer, np.floating)):
            # Handle Excel serial numbers (1899-12-30 base)
            if val > 25569:
                try:
                    return pd.to_datetime(val, unit='D', origin='1899-12-30')
                except:
                    pass
        try:
            return pd.to_datetime(val)
        except:
            return pd.NaT

    def _create_executive_summary(self, wb, df):
        """Sheet 1: Executive Summary"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = 'EXECUTIVE SUMMARY'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='4472C4', fill_type='solid')
        ws.merge_cells('A1:C1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = 'Procurement Analysis Report'
        ws['A2'].font = Font(size=12, bold=True, color='444444')
        ws.merge_cells('A2:C2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Detect Period
        try:
            min_date = pd.to_datetime(df['date']).min().strftime('%Y-%m-%d')
            max_date = pd.to_datetime(df['date']).max().strftime('%Y-%m-%d')
            ws['A3'] = f"Period: {min_date} to {max_date}"
        except:
            ws['A3'] = "Period: Unknown"
        
        ws['A3'].font = Font(italic=True, color='666666')
        ws.merge_cells('A3:C3')
        ws['A3'].alignment = Alignment(horizontal='center')

        # Metrics
        total_spend = df['amount'].sum()
        tx_count = len(df)
        vendor_count = df['vendor'].nunique()
        avg_tx = df['amount'].mean()
        cat_count = df['category'].nunique() if 'category' in df.columns else 0
        
        ws['A5'] = 'KEY METRICS'
        ws['A5'].font = Font(bold=True, underline='single')
        
        metrics = [
            ('Total Spend', f"${total_spend:,.0f}"),
            ('Number of Transactions', f"{tx_count:,}"),
            ('Number of Vendors', f"{vendor_count:,}"),
            ('Average Transaction', f"${avg_tx:,.2f}"),
            ('Number of Categories', f"{cat_count:,}")
        ]
        
        row = 6
        for label, val in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = val
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
            
        # Top 5 Vendors
        row += 2
        ws[f'A{row}'] = "TOP 5 VENDORS BY SPEND"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Table Header
        headers = ['Rank', 'Vendor Name', 'Total Spend', '% of Total', 'Transactions']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            self._apply_header_style(cell)
        
        top_vendors = df.groupby('vendor')['amount'].agg(['sum', 'count']).nlargest(5, 'sum').reset_index()
        top_vendors['pct'] = top_vendors['sum'] / total_spend
        
        row += 1
        for idx, r in top_vendors.iterrows():
            ws.cell(row=row, column=1, value=idx+1)
            ws.cell(row=row, column=2, value=r['vendor'])
            ws.cell(row=row, column=3, value=r['sum']).number_format = '$#,##0'
            ws.cell(row=row, column=4, value=r['pct']).number_format = '0.0%'
            ws.cell(row=row, column=5, value=r['count'])
            row += 1

        # Formatting
        self._apply_standard_formatting(ws)

    def _create_vendor_analysis(self, wb, df):
        """Sheet 2: Spend by Vendor"""
        ws = wb.create_sheet("Spend by Vendor")
        
        # Aggregation
        vendor_analysis = df.groupby('vendor').agg({
            'amount': ['sum', 'count', 'mean']
        }).reset_index()
        
        vendor_analysis.columns = ['Vendor Name', 'Total Spend', 'Transaction Count', 'Average Transaction']
        total_spend = df['amount'].sum()
        vendor_analysis['% of Total Spend'] = vendor_analysis['Total Spend'] / total_spend
        
        # Optional fields
        if 'vendor_country' in df.columns:
            country_map = df.groupby('vendor')['vendor_country'].first()
            vendor_analysis['Vendor Country'] = vendor_analysis['Vendor Name'].map(country_map)
            
        # Sort
        vendor_analysis = vendor_analysis.sort_values('Total Spend', ascending=False).reset_index(drop=True)
        vendor_analysis.insert(0, 'Rank', vendor_analysis.index + 1)
        
        # Rearrange columns if needed
        cols = ['Rank', 'Vendor Name', 'Total Spend', 'Transaction Count', 'Average Transaction', '% of Total Spend']
        if 'Vendor Country' in vendor_analysis.columns:
             cols.append('Vendor Country')
             
        # Write
        self._write_dataframe_to_sheet(ws, vendor_analysis[cols])
        
        # Specific formatting
        for row in ws.iter_rows(min_row=2):
            # Total Spend (Col C)
            row[2].number_format = '$#,##0'
            # Avg Tx (Col E)
            row[4].number_format = '$#,##0.00'
            # Pct (Col F)
            row[5].number_format = '0.0%'
            
            # Top 10 green
            if row[0].value and row[0].value <= 10:
                for cell in row:
                    cell.fill = PatternFill(start_color='E2EFDA', fill_type='solid')

        self._apply_standard_formatting(ws)

    def _create_category_analysis(self, wb, df):
        """Sheet 3: Spend by Category"""
        ws = wb.create_sheet("Spend by Category")
        
        if 'category' not in df.columns:
            ws['A1'] = "No 'category' column available in data."
            return

        category_analysis = df.groupby('category').agg({
            'amount': ['sum', 'count'],
            'vendor': 'nunique'
        }).reset_index()
        
        category_analysis.columns = ['Category', 'Total Spend', 'Transaction Count', 'Number of Vendors']
        category_analysis['% of Total Spend'] = category_analysis['Total Spend'] / df['amount'].sum()
        
        category_analysis = category_analysis.sort_values('Total Spend', ascending=False).reset_index(drop=True)
        category_analysis.insert(0, 'Rank', category_analysis.index + 1)
        
        col_order = ['Rank', 'Category', 'Total Spend', '% of Total Spend', 'Transaction Count', 'Number of Vendors']
        self._write_dataframe_to_sheet(ws, category_analysis[col_order])
        
        # Formatting
        for row in ws.iter_rows(min_row=2):
            row[2].number_format = '$#,##0'
            row[3].number_format = '0.0%'

        self._apply_standard_formatting(ws)

    def _create_monthly_trends(self, wb, df):
        """Sheet 4: Monthly Trends"""
        ws = wb.create_sheet("Monthly Trends")
        
        df_copy = df.copy()
        df_copy['month'] = pd.to_datetime(df_copy['date']).dt.to_period('M')
        
        monthly = df_copy.groupby('month').agg({
            'amount': ['sum', 'count', 'mean'],
            'vendor': 'nunique'
        }).reset_index()
        
        monthly.columns = ['Month', 'Total Spend', 'Transaction Count', 'Average Transaction', 'Number of Vendors']
        monthly['Month'] = monthly['Month'].astype(str) # Convert period to string for Excel
        
        self._write_dataframe_to_sheet(ws, monthly)
        
        for row in ws.iter_rows(min_row=2):
            row[1].number_format = '$#,##0'
            row[3].number_format = '$#,##0.00'

        self._apply_standard_formatting(ws)

    def _create_insights(self, wb, df):
        """Sheet 5: Top Insights"""
        ws = wb.create_sheet("Top Insights")
        
        ws['A1'] = 'TOP INSIGHTS'
        ws['A1'].font = Font(size=14, bold=True, color='4472C4')
        
        insights = []
        
        # 1. Supplier Consolidation
        vendor_stats = df.groupby('vendor')['amount'].sum()
        tail_vendors = vendor_stats[vendor_stats < 5000]
        if not tail_vendors.empty:
            tail_spend = tail_vendors.sum()
            insights.append({
                'title': 'SUPPLIER CONSOLIDATION',
                'finding': f"{len(tail_vendors)} vendors have spend < $5,000",
                'data': f"Total Tail Spend: ${tail_spend:,.0f}",
                'savings': f"Potential Savings: ${tail_spend*0.15:,.0f} - ${tail_spend*0.20:,.0f} (15-20%)",
                'action': "Consolidate to preferred vendors"
            })
            
        # 2. Vendor Concentration
        top_10_spend = vendor_stats.nlargest(10).sum()
        total_spend = df['amount'].sum()
        conc_pct = (top_10_spend / total_spend) * 100
        
        insights.append({
            'title': 'VENDOR CONCENTRATION',
            'finding': f"Top 10 vendors account for {conc_pct:.1f}% of total spend",
            'data': f"Top 10 Spend: ${top_10_spend:,.0f}",
            'savings': "Risk Level: " + ("High" if conc_pct > 80 else "Medium" if conc_pct > 50 else "Low"),
            'action': "Diversify supply base if concentration is high"
        })
        
        current_row = 3
        for idx, insight in enumerate(insights, 1):
            ws[f'A{current_row}'] = f"INSIGHT #{idx}: {insight['title']}"
            ws[f'A{current_row}'].font = Font(bold=True, size=11)
            current_row += 1
            
            ws[f'A{current_row}'] = f"Finding: {insight['finding']}"
            current_row += 1
            ws[f'A{current_row}'] = insight['data']
            current_row += 1
            ws[f'A{current_row}'] = insight['savings']
            current_row += 1
            ws[f'A{current_row}'] = f"Action: {insight['action']}"
            current_row += 2 # Spacer
            
        ws.column_dimensions['A'].width = 80

    def _create_detailed_data(self, wb, df):
        """Sheet 6: Detailed Data"""
        ws = wb.create_sheet("Detailed Data")
        
        # Limit rows for performance if needed, but specification says export entire DF
        # Using dataframe_to_rows is efficient
        if len(df) > 100000:
             # Just a warning or limit could be applied here
             pass

        self._write_dataframe_to_sheet(ws, df)
        self._apply_standard_formatting(ws)

    def _create_data_quality_report(self, wb, df):
        """Sheet 7: Data Quality Report"""
        ws = wb.create_sheet("Data Quality Report")
        
        ws['A1'] = 'DATA QUALITY ASSESSMENT'
        ws['A1'].font = Font(size=14, bold=True)
        
        completeness = (df.notna().sum() / len(df)) * 100
        overall_score = completeness.mean()
        
        ws['A2'] = f"Overall Quality Score: {overall_score:.1f}%"
        ws['A2'].font = Font(bold=True)
        
        # Table Header
        ws['A4'] = 'Field Name'
        ws['B4'] = 'Completeness %'
        ws['C4'] = 'Status'
        
        for cell in ws['4:4']:
            if cell.value:
                self._apply_header_style(cell)
        
        row = 5
        for col in df.columns:
            comp_val = completeness[col]
            ws[f'A{row}'] = col
            ws[f'B{row}'] = comp_val / 100 # Store as decimal for % formatting
            ws[f'B{row}'].number_format = '0.0%'
            
            status_cell = ws[f'C{row}']
            if comp_val >= 90:
                status_cell.value = 'OK'
                status_cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid') # Green
            elif comp_val >= 70:
                status_cell.value = 'WARNING'
                status_cell.fill = PatternFill(start_color='FFE699', fill_type='solid') # Yellow
            else:
                status_cell.value = 'CRITICAL'
                status_cell.fill = PatternFill(start_color='FFC7CE', fill_type='solid') # Red
                
            row += 1
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _write_dataframe_to_sheet(self, ws, df):
        """Helper to write DF with headers"""
        rows = dataframe_to_rows(df, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    self._apply_header_style(cell)
                    
        # Freeze panes
        ws.freeze_panes = 'A2'
        # Auto Filter
        ws.auto_filter.ref = ws.dimensions

    def _apply_header_style(self, cell):
        """Blue header, white text"""
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_standard_formatting(self, ws):
        """Apply borders and auto-width"""
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                
        # Simple auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Usage Example (if run directly)
if __name__ == "__main__":
    # Create dummy data
    data = {
        'vendor': ['Vendor A', 'Vendor B', 'Vendor A', 'Vendor C', 'Vendor B'] * 100,
        'amount': np.random.uniform(100, 10000, 500),
        'date': pd.date_range(start='2023-01-01', periods=500),
        'category': ['IT', 'Services', 'IT', 'Hardware', 'Services'] * 100
    }
    df = pd.DataFrame(data)
    
    loader = DataLoader()
    path = loader.load(df, output_path='test_output.xlsx')
    print(f"File generated: {path}")

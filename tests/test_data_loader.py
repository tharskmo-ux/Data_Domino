import pytest
import pandas as pd
import numpy as np
import os
from data_loader import DataLoader
from openpyxl import load_workbook

@pytest.fixture
def sample_df():
    data = {
        'vendor': ['Vendor A', 'Vendor B', 'Vendor A'],
        'amount': [100.0, 200.0, 150.0],
        'date': ['2023-01-01', '2023-01-02', '2023-01-03'],
        'category': ['IT', 'Services', 'IT']
    }
    return pd.DataFrame(data)

def test_basic_load(sample_df, tmp_path):
    output_file = tmp_path / "report.xlsx"
    loader = DataLoader()
    path = loader.load(sample_df, output_path=str(output_file))

    assert os.path.exists(path)
    wb = load_workbook(path)
    assert 'Executive Summary' in wb.sheetnames
    assert 'Spend by Vendor' in wb.sheetnames
    assert 'Detailed Data' in wb.sheetnames

def test_empty_df():
    loader = DataLoader()
    with pytest.raises(ValueError, match="Cannot generate report from empty DataFrame"):
        loader.load(pd.DataFrame())

def test_missing_required_fields():
    loader = DataLoader()
    df = pd.DataFrame({'vendor': ['A'], 'amount': [10]})
    # Now it should raise Missing required fields: ['date']
    with pytest.raises(ValueError, match="Missing required fields: \['date'\]"):
        loader.load(df)

def test_mapping_info_usage(tmp_path):
    # This test verifies that mapping_info now works
    output_file = tmp_path / "mapped_report.xlsx"
    df = pd.DataFrame({
        'Supplier': ['Vendor A'],
        'Cost': [100.0],
        'Transaction Date': ['2023-01-01']
    })
    mapping = {
        'vendor': 'Supplier',
        'amount': 'Cost',
        'date': 'Transaction Date'
    }
    loader = DataLoader()
    path = loader.load(df, mapping_info=mapping, output_path=str(output_file))

    assert os.path.exists(path)
    wb = load_workbook(path)
    ws = wb['Executive Summary']
    assert ws['B6'].value == '$100'

def test_excel_serial_dates(tmp_path):
    df = pd.DataFrame({
        'vendor': ['Vendor A'],
        'amount': [100.0],
        'date': [44562] # 2022-01-01
    })
    output_file = tmp_path / "excel_dates.xlsx"
    loader = DataLoader()
    path = loader.load(df, output_path=str(output_file))

    wb = load_workbook(path)
    ws = wb['Executive Summary']
    # Period: 2022-01-01 to 2022-01-01
    assert "2022-01-01" in ws['A3'].value

def test_data_quality_report(tmp_path):
    df = pd.DataFrame({
        'vendor': ['Vendor A', None, 'Vendor C'],
        'amount': [100.0, 200.0, None],
        'date': ['2023-01-01', '2023-01-02', '2023-01-03']
    })
    output_file = tmp_path / "quality_report.xlsx"
    loader = DataLoader()
    path = loader.load(df, output_path=str(output_file))

    wb = load_workbook(path)
    assert 'Data Quality Report' in wb.sheetnames
    ws = wb['Data Quality Report']

    assert ws['A5'].value == 'vendor'
    assert ws['B5'].value == pytest.approx(0.666, 0.01)
    # 66.6% is < 70%, so CRITICAL
    assert ws['C5'].value == 'CRITICAL'

    assert ws['A7'].value == 'date'
    assert ws['B7'].value == 1.0
    assert ws['C7'].value == 'OK'

def test_calculations_accuracy(sample_df, tmp_path):
    output_file = tmp_path / "calc_test.xlsx"
    loader = DataLoader()
    path = loader.load(sample_df, output_path=str(output_file))

    wb = load_workbook(path)
    ws_summary = wb['Executive Summary']

    # Total Spend should be 100 + 200 + 150 = 450
    assert ws_summary['B6'].value == '$450'
    # Num Transactions should be 3
    assert ws_summary['B7'].value == '3'
    # Num Vendors should be 2
    assert ws_summary['B8'].value == '2'

    ws_vendor = wb['Spend by Vendor']
    # Check Vendor A spend: 100 + 150 = 250
    # Vendor A should be Rank 1
    assert ws_vendor['B2'].value == 'Vendor A'
    assert ws_vendor['C2'].value == 250
